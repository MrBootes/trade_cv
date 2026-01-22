from __future__ import annotations

from typing import Optional

from datetime import datetime
import os
import sys
import time

import pandas as pd
from openpyxl import load_workbook

from loaders.load_validate import identify_trade_type, read_input_file


def _normalize_mode(mode: str | None) -> str | None:
	if mode is None:
		return None
	m = str(mode).strip().upper()
	if m in {"UNO", "ONE", "ONES", "ONE-SIDED", "ONE_SIDED", "ONE SIDED", "TYPE1", "TYPE 1", "1"}:
		return "ONES"
	if m in {"INOUT", "TWO", "TWOS", "TWO-SIDED", "TWO_SIDED", "TWO SIDED", "TYPE2", "TYPE 2", "2"}:
		return "TWOS"
	if m in {"AUTO", "A", "3", "TYPE3", "TYPE 3", ""}:
		return "AUTO"
	return m


def _excel_col_letter(idx0: int) -> str:
	idx = int(idx0) + 1
	out = []
	while idx > 0:
		idx, rem = divmod(idx - 1, 26)
		out.append(chr(ord('A') + rem))
	return ''.join(reversed(out))


def _excel_letter_to_idx0(s: str) -> int | None:
	if not s:
		return None
	s = str(s).strip().upper()
	if not s.isalpha():
		return None
	idx = 0
	for ch in s:
		idx = idx * 26 + (ord(ch) - ord('A') + 1)
	return idx - 1


def _preview_columns(df: pd.DataFrame, *, max_cols: int = 60) -> None:
	if df is None or not isinstance(df, pd.DataFrame):
		print("No DataFrame to preview")
		return
	cols = list(df.columns)
	print("\nAvailable columns:")
	for i, col in enumerate(cols[:max_cols], start=1):
		series = df[col] if col in df.columns else None
		sample = ""
		try:
			if series is not None:
				non_null = series.dropna()
				if not non_null.empty:
					sample = str(non_null.iloc[0])
		except Exception:
			sample = ""
		letter = _excel_col_letter(i - 1)
		if sample:
			print(f"  {i:>2}) {letter}: {col}   (sample: {sample})")
		else:
			print(f"  {i:>2}) {letter}: {col}")
	if len(cols) > max_cols:
		print(f"  ... and {len(cols) - max_cols} more column(s)")


def _select_column(df: pd.DataFrame, prompt: str, *, allow_blank: bool) -> str | None:
	cols = list(df.columns)
	while True:
		ans = input(prompt).strip().strip('"').strip("'")
		if ans == "":
			return None if allow_blank else None

		if ans.isdigit():
			idx1 = int(ans)
			if 1 <= idx1 <= len(cols):
				return str(cols[idx1 - 1])
			print(f"Invalid index: {idx1}. Expected 1..{len(cols)}")
			continue

		idx0 = _excel_letter_to_idx0(ans)
		if idx0 is not None:
			if 0 <= idx0 < len(cols):
				return str(cols[idx0])
			print(f"Invalid column letter: {ans}. Expected A..{_excel_col_letter(len(cols)-1)}")
			continue

		if ans in df.columns:
			return ans

		lower_map = {str(c).strip().lower(): c for c in df.columns}
		key = ans.strip().lower()
		if key in lower_map:
			return str(lower_map[key])

		print(f"Column '{ans}' not found. Enter index, letter, or exact column name.")


def _prompt_manual_mapping(df: pd.DataFrame, mode: str, *, interactive: bool) -> dict | None:
	if not interactive:
		return None
	if df is None or not isinstance(df, pd.DataFrame):
		return None
	mode = _normalize_mode(mode)
	if mode not in {"ONES", "TWOS"}:
		return None

	ans = input("\nDo you want to enter column mapping manually (index/letter/name)? (y/n): ").strip().lower()
	if ans not in {"y", "yes"}:
		return None

	_preview_columns(df)
	print("\nEnter column by: number (1..N), Excel letter (A..), or exact name. Blank = auto-detect.")

	mapping: dict[str, str] = {}
	if mode == "ONES":
		fields = [
		 ("boards", True),
		 ("type", False),
		 ("tickers", False),
		 ("date_trade", False),
		 ("volume_signed", False),
		 ("price_trade", False),
		 ("commission", True),
		]
		for field, optional in fields:
			col = _select_column(df, f"  {field}{' (optional)' if optional else ''}: ", allow_blank=True)
			if col:
				mapping[field] = col

	elif mode == "TWOS":
		fields = [
		 ("boards", True),
		 ("type", False),
		 ("tickers", False),
		 ("volume", False),
		 ("dates_buy", False),
		 ("buy", False),
		 ("buy_commission", True),
		 ("dates_sell", False),
		 ("sell", False),
		 ("sell_commission", True),
		]
		for field, optional in fields:
			col = _select_column(df, f"  {field}{' (optional)' if optional else ''}: ", allow_blank=True)
			if col:
				mapping[field] = col

	return mapping or None


def _is_missing_scalar(value) -> bool:
	if value is None:
		return True
	try:
		return bool(pd.isna(value))
	except Exception:
		return False


def _num_to_clean_str(x) -> str:
	if x is None:
		return ""

	try:
		import pandas as _pd
		if _pd.isna(x):
			return ""
	except Exception:
		pass
	from decimal import Decimal
	if isinstance(x, Decimal):
		return format(x, 'f')
	if isinstance(x, bool):
		return "1" if x else "0"
	if isinstance(x, int):
		return str(x)
	if isinstance(x, float):

		s = f"{x:.12f}".rstrip('0').rstrip('.')
		return s if s else "0"
	return str(x).strip()


def _uno_to_fifo_rows(data: dict):
	if not isinstance(data, dict):
		raise ValueError("data must be a dict")

	boards = data.get('boards', []) or []
	types = data.get('type', []) or []
	tickers = data.get('tickers', []) or []
	dates = data.get('date_trade', []) or []
	vols = data.get('volume_signed', []) or []
	prices = data.get('price_trade', []) or []
	comms = data.get('commission', []) or []

	n = max(len(tickers), len(dates), len(vols), len(prices), len(types), len(boards), len(comms))
	if n == 0:
		return [], False, False

	def _get(lst, i):
		return lst[i] if i < len(lst) else None

	rows = []
	has_board = False
	has_comm = False

	from file_managers import fifo_match as fifo_pkg

	for i in range(n):
		typ = _get(types, i)
		tkr = _get(tickers, i)
		dt = _get(dates, i)
		vol = _get(vols, i)
		price = _get(prices, i)
		board = _get(boards, i)
		comm = _get(comms, i)

		if typ is None and tkr is None and dt is None and vol is None and price is None and board is None and comm is None:
			continue

		board_s = (str(board).strip() if board is not None else "")
		if board_s:
			has_board = True

		date_obj = None
		if dt is not None:
			try:
				if hasattr(dt, 'to_pydatetime'):
					dt = dt.to_pydatetime()
			except Exception:
				pass
			try:
				if hasattr(dt, 'date'):
					date_obj = dt.date() if isinstance(dt, datetime) else dt
				else:
					date_obj = dt
			except Exception:
				date_obj = dt

		if vol is None:
			vol_i = 0
		else:
			try:
				vol_f = float(vol)

				if abs(vol_f - round(vol_f)) > 1e-9:
					raise ValueError(f"Row {i + 1}: volume_signed must be an integer, got {vol}")
				vol_i = int(round(vol_f))
			except Exception as e:
				raise ValueError(f"Row {i + 1}: volume_signed='{vol}' is invalid") from e

		price_raw = _num_to_clean_str(price)
		if not price_raw and vol_i != 0:
			raise ValueError(f"Row {i + 1}: price_trade is required for non-zero volume")
		p_dec, p_shown, p_frac = fifo_pkg.parse_price_keep_format(price_raw) if price_raw else (None, "", 0)

		c_dec = None
		c_shown = ""
		c_frac = 0
		comm_raw = _num_to_clean_str(comm)
		if comm_raw:
			has_comm = True
			c_dec, c_shown, c_frac = fifo_pkg.parse_price_keep_format(comm_raw)

		rows.append({
		 "TYPE": (str(typ).strip() if typ is not None else ""),
		 "TICKER": (str(tkr).strip() if tkr is not None else ""),
		 "DATE": date_obj,
		 "VOLUME": vol_i,
		 "PRICE_DEC": p_dec,
		 "PRICE_STR": p_shown,
		 "PRICE_FRAC": p_frac,
		 "COMM_DEC": c_dec,
		 "COMM_STR": c_shown,
		 "COMM_FRAC": c_frac,
		 "MARKETBOARD": board_s,
		 "_ord": i,
		})

	rows.sort(key=lambda x: (x.get("TICKER", ""), x.get("DATE"), x.get("_ord", 0)))
	return rows, has_comm, has_board


def _data_to_dataframe(data: dict) -> pd.DataFrame:
	if not isinstance(data, dict):
		raise ValueError("data must be a dict")
	visible = {
	 k: v
	 for k, v in data.items()
	 if not str(k).startswith('_') and isinstance(v, list)
	}
	if not visible:
		return pd.DataFrame()


	lengths = {k: len(v) for k, v in visible.items()}
	if len(set(lengths.values())) > 1:
		raise ValueError(f"Cannot save: columns have different lengths: {lengths}")

	return pd.DataFrame(visible)





def _prepare_dataframe_for_saving(df_out: pd.DataFrame, *, fmt: str) -> pd.DataFrame:
	if df_out is None or df_out.empty:
		return df_out

	df = df_out.copy()



	for col in df.columns:
		if 'date' not in str(col).lower():
			continue
		series = df[col]
		if pd.api.types.is_datetime64_any_dtype(series):
			continue
		if pd.api.types.is_numeric_dtype(series):

			non_null = series.dropna()
			if not non_null.empty:
				try:
					as_float = pd.to_numeric(non_null, errors='coerce')
					as_float = as_float.dropna()
					if not as_float.empty:
						mn = float(as_float.min())
						mx = float(as_float.max())

						if 20000 <= mn <= 80000 and 20000 <= mx <= 80000:
							df[col] = pd.to_datetime(series, unit='D', origin='1899-12-30', errors='coerce')
				except Exception:
					pass
		else:

			parsed = pd.to_datetime(series, errors='coerce')
			if parsed.notna().any():
				df[col] = parsed

	if fmt == 'csv':
		for col in df.columns:
			if 'date' in str(col).lower() and pd.api.types.is_datetime64_any_dtype(df[col]):

				try:
					series = df[col]
					is_midnight = (series.dt.hour == 0) & (series.dt.minute == 0) & (series.dt.second == 0)
					if bool(is_midnight.all()):
						df[col] = series.dt.strftime('%Y-%m-%d')
					else:
						df[col] = series.dt.strftime('%Y-%m-%d %H:%M:%S')
				except Exception:

					df[col] = df[col].astype(str)
	return df


def _offer_save_output(data: dict, *, mode: str, drop_columns: set[str] | None = None) -> None:
	try:
		df_out = _data_to_dataframe(data)
	except Exception as e:
		print(f"\n⚠ Could not prepare data for saving: {e}")
		return

	if drop_columns:
		try:
			df_out = df_out.drop(columns=[c for c in drop_columns], errors='ignore')
		except Exception:
			pass

	if df_out.empty:
		return

	answer = input("\nSave validated/sorted data to a file in this program folder? (y/n): ").strip().lower()
	if answer != 'y':
		return

	print("Choose format:")
	print("  1) CSV")
	print("  2) Excel (.xlsx)")
	fmt_choice = input("> ").strip().lower()
	fmt = 'csv' if fmt_choice in {'', '1', 'csv'} else 'xlsx'

	stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
	default_name = f"validated_sorted_{mode.lower()}_{stamp}.{fmt}"
	name = input(f"Enter output filename (blank for '{default_name}'): ").strip().strip('"')
	filename = name or default_name

	program_dir = os.path.dirname(os.path.abspath(__file__))
	out_path = os.path.join(program_dir, filename)

	def _apply_excel_date_formats(path: str) -> None:

		last_err: Exception | None = None
		for _ in range(6):
			try:
				wb = load_workbook(path)
				ws = wb.active
				headers = [cell.value for cell in ws[1]]
				date_cols = []
				for idx, name in enumerate(headers, start=1):
					if name is None:
						continue
					if 'date' in str(name).lower():
						date_cols.append(idx)

				for col_idx in date_cols:
					for row_idx in range(2, ws.max_row + 1):
						cell = ws.cell(row=row_idx, column=col_idx)
						if cell.value is None:
							continue
						cell.number_format = 'yyyy-mm-dd'

				wb.save(path)
				return
			except Exception as e:
				last_err = e
				time.sleep(0.2)
		if last_err is not None:
			raise last_err

	try:
		df_save = _prepare_dataframe_for_saving(df_out, fmt=fmt)
		if fmt == 'xlsx':

			with pd.ExcelWriter(
			 out_path,
			 engine='openpyxl',
			 date_format='yyyy-mm-dd',
			 datetime_format='yyyy-mm-dd',
			) as writer:
				df_save.to_excel(writer, index=False)


			try:
				_apply_excel_date_formats(out_path)
			except Exception as e:
				print(f"\n⚠ Saved XLSX, but could not apply date formatting: {e}")
		else:

			df_save.to_csv(out_path, index=False, encoding='utf-8-sig')
		print(f"✓ Saved: {out_path}")
	except Exception as e:
		print(f"\nError saving file: {e}")


def _prompt_file_path() -> Optional[str]:
	file_path = input("\nEnter the path to your Excel or CSV file (blank/none/skip to skip): ").strip().strip('"')
	if file_path == "" or file_path.lower() in {"none", "skip", "null"}:
		return None
	return file_path


def _prompt_sheet_name() -> Optional[str]:
	sheet = input("Enter Excel sheet name (blank for default / for CSV): ").strip().strip('"')
	return sheet or None


def _prompt_load_mode() -> str:
	print("\nChoose load mode:")
	print("  1) Type 1 (ONES / one-sided)")
	print("  2) Type 2 (TWOS / two-sided)")
	print("  3) Auto")
	choice = input("> ").strip().lower()
	if choice in {"1", "type1", "uno", "ones", "one", "one-sided", "one_sided", "one sided"}:
		return "ONES"
	if choice in {"2", "type2", "inout", "twos", "two", "two-sided", "two_sided", "two sided"}:
		return "TWOS"
	if choice in {"3", "auto", "a", ""}:
		return "AUTO"
	raise ValueError("Invalid choice. Use 1, 2, or 3 (Auto).")


def _prompt_optional_flow_file_path(default_path: Optional[str]) -> Optional[str]:
	ans = input(
	 "\nEnter inflow/outflow data file path (blank/none/skip = skip, 'same' = use same file): "
	).strip().strip('"')
	if ans == "" or ans.lower() in {"skip", "none", "null"}:
		return None
	if ans.lower() == "same":
		return default_path
	return ans


def _coerce_trades_to_dict(trade_data):
	if trade_data is None:
		return None

	def _is_missing(v) -> bool:
		return _is_missing_scalar(v)

	def _normalize_missing_in_dict(d: dict) -> dict:
		out = dict(d)
		for k, v in d.items():
			if not isinstance(v, list):
				continue
			out[k] = [None if _is_missing(x) else x for x in v]
		return out

	def _normalize_dict_only(d: dict) -> dict:
	 # IMPORTANT: Do not auto-add or auto-fill marketboard here.
		return _normalize_missing_in_dict(d)

	def _to_float(x):
		if _is_missing(x):
			return None
		try:
			return float(x)
		except Exception:
			return None

	def _parse_fifo_date(s):
		if _is_missing(s):
			return None
		if hasattr(s, 'year') and hasattr(s, 'month') and hasattr(s, 'day') and not isinstance(s, str):
			return s
		s = str(s).strip()
		if not s:
			return None
		try:
			return datetime.strptime(s, "%d.%m.%Y")
		except Exception:
			try:
				return datetime.strptime(s, "%Y-%m-%d")
			except Exception:
				return None


	if isinstance(trade_data, dict):
		return _normalize_dict_only(trade_data)


	if isinstance(trade_data, list):
		if len(trade_data) == 0:
			return {
			 'boards': [],
			 'type': [],
			 'tickers': [],
			 'volume': [],
			 'dates_buy': [],
			 'buy': [],
			 'dates_sell': [],
			 'sell': [],
			}
		if not all(isinstance(r, dict) for r in trade_data):
			raise TypeError("trade_data list must contain dict rows")

		has_commission = any(('_BC_NUM' in r) or ('_SC_NUM' in r) or ('BUY_COMMISSION' in r) or ('SELL_COMMISSION' in r) for r in trade_data)

		rows = []
		for r in trade_data:
			ticker = (r.get('TICKER') or r.get('ticker') or '')
			ticker = '' if _is_missing(ticker) else str(ticker).strip()
			typ = (r.get('TYPE') or r.get('type') or '')
			typ = '' if _is_missing(typ) else str(typ).strip()
			board = (r.get('MARKETBOARD') or r.get('board') or r.get('BOARD') or None)
			board = None if _is_missing(board) or str(board).strip() == '' else str(board).strip()
			vol = _to_float(r.get('VOLUME'))
			if vol is None:
				vol = 0.0
			buy_num = r.get('_PB_NUM') if '_PB_NUM' in r else r.get('PRICE_BUY_NUM')
			sell_num = r.get('_PS_NUM') if '_PS_NUM' in r else r.get('PRICE_SELL_NUM')
			row = {
			 'boards': board,
			 'type': typ,
			 'tickers': ticker,
			 'volume': vol,
			 'dates_buy': _parse_fifo_date(r.get('DATE_BUY')),
			 'buy': _to_float(buy_num),
			 'dates_sell': _parse_fifo_date(r.get('DATE_SELL')),
			 'sell': _to_float(sell_num),
			}
			if has_commission:
				bc_num = r.get('_BC_NUM') if '_BC_NUM' in r else r.get('BUY_COMM_NUM')
				sc_num = r.get('_SC_NUM') if '_SC_NUM' in r else r.get('SELL_COMM_NUM')
				row['buy_commission'] = _to_float(bc_num)
				row['sell_commission'] = _to_float(sc_num)
			rows.append(row)

		def _date_sort_key(x):

			return (1, None) if x is None else (0, x)

		rows.sort(
		 key=lambda r: (
		  r.get('boards') or '',
		  r.get('tickers') or '',
		  _date_sort_key(r.get('dates_buy')),
		  _date_sort_key(r.get('dates_sell')),
		 ),
		)

		out = {
		 'boards': [r['boards'] for r in rows],
		 'type': [r['type'] for r in rows],
		 'tickers': [r['tickers'] for r in rows],
		 'volume': [r['volume'] for r in rows],
		 'dates_buy': [r['dates_buy'] for r in rows],
		 'buy': [r['buy'] for r in rows],
		 'dates_sell': [r['dates_sell'] for r in rows],
		 'sell': [r['sell'] for r in rows],
		}
		if has_commission:
			out['buy_commission'] = [r.get('buy_commission') for r in rows]
			out['sell_commission'] = [r.get('sell_commission') for r in rows]
		return out

	raise TypeError(f"Unsupported trade_data type: {type(trade_data).__name__}")


def main():
	print("\nTrade Loader Starter")
	print("=" * 50)

	loaded_data = [None, None]

	file_path = _prompt_file_path()
	loader_interactive = sys.stdin.isatty()
	primary_data = None

	if file_path is None:
		print("\nSkipping main trade load; proceeding to FLOW step...")
		sheet_name = None
		mode = None
		df = None
	else:
		sheet_name = _prompt_sheet_name()
		mode = _prompt_load_mode()

		df = read_input_file(file_path, sheet_name=sheet_name)

	if mode == "AUTO":
		detected = identify_trade_type(df, interactive=False)
		print(f"\nAuto-detected type: {detected}")
		mode = detected

	mode = _normalize_mode(mode)

	custom_column_names = _prompt_manual_mapping(df, mode, interactive=loader_interactive) if df is not None else None

	primary_data = None
	if mode == "ONES":
		from loaders.ones_load import display_data, read_trade_data

		primary_data = read_trade_data(
		 file_path,
		 sheet_name=sheet_name,
		 df=df,
		 interactive=loader_interactive,
		 custom_column_names=custom_column_names,
		)
		if primary_data:
			original_has_board = bool(primary_data.get('_original_has_board', False))
			default_board = None
			boards = primary_data.get('boards')
			if isinstance(boards, list):
				for b in boards:
					if not _is_missing_scalar(b) and str(b).strip():
						default_board = str(b).strip()
						break
			display_data(primary_data)
			_offer_save_output(primary_data, mode=mode, drop_columns=({'boards'} if not original_has_board else None))
			try:
				uno_rows, has_comm, has_board = _uno_to_fifo_rows(primary_data)
				from file_managers import fifo_match as fifo_pkg


				fifo_rows_raw = fifo_pkg.fifo_match(uno_rows, has_commission=has_comm, has_marketboard=original_has_board)
				print(f"\n✓ FIFO matched rows: {len(fifo_rows_raw)}")
				fifo_rows = _coerce_trades_to_dict(fifo_rows_raw)


				if not original_has_board and default_board is not None:
					nn = len(fifo_rows.get('tickers', []) or [])
					fifo_rows = dict(fifo_rows)
					fifo_rows['boards'] = [default_board for _ in range(nn)]
				drop_cols = {'boards'} if not original_has_board else None
				_offer_save_output(fifo_rows, mode="FIFO", drop_columns=drop_cols)
				loaded_data[0] = fifo_rows
			except Exception as e:
				print(f"\n⚠ FIFO match skipped due to error: {e}")

	if mode == "TWOS":
		from loaders.twos_load import display_data, read_trade_data

		primary_data = read_trade_data(
		 file_path,
		 sheet_name=sheet_name,
		 df=df,
		 interactive=loader_interactive,
		 custom_column_names=custom_column_names,
		)
		if primary_data:
			original_has_board = bool(primary_data.get('_original_has_board', False))

			normalized = _coerce_trades_to_dict(primary_data)
			display_data(normalized)
			_offer_save_output(normalized, mode=mode, drop_columns=({'boards'} if not original_has_board else None))
			loaded_data[0] = normalized

	if mode is not None and mode not in {"ONES", "TWOS"}:
		loaded_data[0] = None
		raise RuntimeError(f"Unexpected mode: {mode}")


	try:
		flow_path = _prompt_optional_flow_file_path(file_path)
		if flow_path:

			flow_sheet = sheet_name
			if file_path is None or os.path.abspath(flow_path) != os.path.abspath(file_path):
				flow_sheet = input("Enter Excel sheet name for FLOW (blank for default / for CSV): ").strip().strip('"') or None

			from loaders.flow_load import display_data as flow_display, read_trade_data as flow_read
			flow_data = flow_read(flow_path, sheet_name=flow_sheet, interactive=loader_interactive)
			if flow_data:
				flow_display(flow_data)
				_offer_save_output(flow_data, mode="FLOW")
				loaded_data[1] = flow_data
	except Exception as e:
		print(f"\n⚠ FLOW load skipped due to error: {e}")

	return loaded_data[0], loaded_data[1]

if __name__ == "__main__":
	main()
