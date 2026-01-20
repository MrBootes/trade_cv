"""Helpers for converting broker exports into the internal trade format."""

import csv
import difflib
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, getcontext
from pathlib import Path

getcontext().prec = 50

DATE_FMT = "%d.%m.%Y"


def _try_parse_date_multi(s: str):
	txt = (s or "").strip()
	if not txt:
		raise ValueError("empty date")
	for fmt in (DATE_FMT, "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
		try:
			return datetime.strptime(txt, fmt).date()
		except Exception:
			pass

	try:
		return datetime.fromisoformat(txt.replace("Z", "+00:00")).date()
	except Exception:
		pass
	raise ValueError(f"invalid date: '{s}'")


def parse_date(s: str):
	return _try_parse_date_multi(s)


def _norm_col_name(name: str) -> str:
	if name is None:
		return ""
	s = str(name).strip().casefold()

	for ch in ("\u00A0", "\u202F", "\u2009", " ", "\t", "\n", "\r"):
		s = s.replace(ch, "")

	s = "".join(c for c in s if c.isalnum())
	return s


def _to_int_loose(v) -> int:
	if v is None:
		return 0
	if isinstance(v, bool):
		return int(v)
	if isinstance(v, int):
		return v
	if isinstance(v, float):
		return int(v)
	if isinstance(v, Decimal):
		return int(v)
	txt = str(v).strip()
	if not txt:
		return 0

	txt = strip_space_digits(txt)

	txt = txt.replace(",", "")
	try:
		return int(txt)
	except Exception:
		try:
			return int(float(txt))
		except Exception:
			return 0


def _resolve_columns(
 header: list[str],
 required: dict[str, list[str]],
 optional: dict[str, list[str]] | None = None,
 *,
 fuzzy_threshold: float = 0.86,
):
	optional = optional or {}

	norm_to_orig: dict[str, str] = {}
	for h in header:
		n = _norm_col_name(h)
		if n and n not in norm_to_orig:
			norm_to_orig[n] = h

	def candidates_for(canon: str) -> list[str]:
		alts = [canon] + list(required.get(canon, [])) + list(optional.get(canon, []))

		seen = set()
		out = []
		for a in alts:
			n = _norm_col_name(a)
			if n and n not in seen:
				seen.add(n)
				out.append(n)
		return out

	resolved: dict[str, str] = {}
	missing: list[str] = []

	all_norm_headers = list(norm_to_orig.keys())
	for canon, syns in required.items():

		found = None
		for alt_norm in candidates_for(canon):
			if alt_norm in norm_to_orig:
				found = norm_to_orig[alt_norm]
				break

		if found is None and all_norm_headers:
			best_norm = None
			best_score = 0.0

			acceptable = candidates_for(canon)
			for hn in all_norm_headers:
				for an in acceptable:
					sc = difflib.SequenceMatcher(a=hn, b=an).ratio()
					if sc > best_score:
						best_score = sc
						best_norm = hn
			if best_norm is not None and best_score >= fuzzy_threshold:
				found = norm_to_orig[best_norm]
		if found is None:
			missing.append(canon)
		else:
			resolved[canon] = found


	for canon, syns in optional.items():
		for alt_norm in candidates_for(canon):
			if alt_norm in norm_to_orig:
				resolved[canon] = norm_to_orig[alt_norm]
				break

	if missing:

		avail = [h for h in header if (h or "").strip()]
		suggestions = {}
		for canon in missing:

			expected_norms = candidates_for(canon)

			pairs = []
			for h in avail:
				hn = _norm_col_name(h)
				if not hn:
					continue
				best = max((difflib.SequenceMatcher(a=hn, b=en).ratio() for en in expected_norms), default=0.0)
				pairs.append((best, h))
			pairs.sort(reverse=True)
			suggestions[canon] = [h for sc, h in pairs[:5] if sc >= 0.55]

		msg_lines = [
		 f"Missing required columns: {missing}.",
		 f"Found columns: {avail}",
		]
		for canon in missing:
			if suggestions.get(canon):
				msg_lines.append(f"Suggestions for '{canon}': {suggestions[canon]}")
		raise ValueError("\n".join(msg_lines))

	return resolved



PRICE_RE = re.compile(
 r"""
	^\s*
	(?P<sign>[+-]?)\s*
	(?P<int>[0-9\u00A0\u202F\u2009\s]*[0-9])?
	(?:[,.](?P<frac>[0-9]+))?
	.*?$
""",
 re.X,
)


def strip_space_digits(s: str) -> str:
	return (
	 (s or "")
	 .replace("\u00A0", "")
	 .replace("\u202F", "")
	 .replace("\u2009", "")
	 .replace(" ", "")
	)


def group_thousands(int_digits: str) -> str:
	s = int_digits
	out = []
	while s:
		out.append(s[-3:])
		s = s[:-3]
	return " ".join(reversed(out)) if out else "0"


def parse_price_keep_format(raw: str):
	txt = (raw or "").strip()

	txt = (
	 txt.replace("руб.", "")
	 .replace("р.", "")
	 .replace("₽", "")
	 .replace("$", "")
	 .replace("€", "")
	 .replace("£", "")
	 .replace("¥", "")
	)
	m = PRICE_RE.match(txt)
	if not m:
		raise ValueError(f"invalid price: '{raw}'")
	sign = m.group("sign") or ""
	int_part = strip_space_digits(m.group("int") or "0")
	frac_part = m.group("frac")
	if not int_part.isdigit():
		raise ValueError(f"invalid integer part of price: '{raw}'")

	dec_str = f"{'-' if sign == '-' else ''}{int_part}"
	frac_len = 0
	if frac_part is not None:
		dec_str += "." + frac_part
		frac_len = len(frac_part)
	try:
		dec_val = Decimal(dec_str)
	except InvalidOperation as e:
		raise ValueError(f"failed to parse price: '{raw}' -> '{dec_str}'") from e

	shown_int = group_thousands(int_part.lstrip("0") or "0")
	if frac_part is None:
		shown = f"{shown_int} ₽"
	else:
		shown = f"{shown_int},{frac_part} ₽"
	if sign == "-":
		shown = "-" + shown
	return dec_val, shown, frac_len


def format_num_for_csv(dec: Decimal | None, frac_len: int) -> str:
	if dec is None:
		return ""
	frac_len = int(frac_len or 0)
	if frac_len <= 0:
		return f"{dec.quantize(Decimal('1'))}"
	q = Decimal(1).scaleb(-frac_len)
	return f"{dec.quantize(q)}"


def sniff_delimiter(path: Path) -> str:
	text = path.read_text(encoding="utf-8-sig", errors="ignore")
	first_lines = [l for l in text.splitlines()[:5] if l.strip()]
	candidates = [";", "\t", ","]
	best, best_cols = ",", 0
	for cand in candidates:
		cols = min((len(l.split(cand)) for l in first_lines), default=0)
		if cols > best_cols:
			best_cols, best = cols, cand
	try:
		sniff = csv.Sniffer().sniff(text[:4096], delimiters=";,\t")
		sn_cols = len(first_lines[0].split(sniff.delimiter)) if first_lines else 0
		if sn_cols >= best_cols:
			return sniff.delimiter
	except Exception:
		pass
	return best



MKT_COL = "MARKETBOARD"
INOUT_BASE_COLS = ["TYPE", "TICKER", "VOLUME", "DATE_BUY", "PRICE_BUY", "DATE_SELL", "PRICE_SELL"]


_INOUT_REQUIRED_SYNONYMS = {
 "TYPE": ["type", "typ", "assettype", "вид", "тип"],
 "TICKER": ["ticker", "symbol", "security", "sec", "тикер", "тик", "бумага", "инструмент"],
 "VOLUME": ["volume", "qty", "quantity", "amount", "count", "количество", "колво", "кол-во", "шт", "units"],
 "DATE_BUY": ["datebuy", "buydate", "buy_date", "date_buy", "datepurchase", "purchasedate", "датыпокупки", "датапокупки", "дата покупки", "дата_покупки"],
 "PRICE_BUY": ["pricebuy", "buyprice", "buy_price", "price_buy", "purchaseprice", "buy", "цена покупки", "ценапокупки", "цена_покупки"],
 "DATE_SELL": ["datesell", "selldate", "sell_date", "date_sell", "datesale", "saledate", "датапродажи", "дата продажи", "дата_продажи"],
 "PRICE_SELL": ["pricesell", "sellprice", "sell_price", "price_sell", "saleprice", "sell", "цена продажи", "ценапродажи", "цена_продажи"],
}


_INOUT_OPTIONAL_SYNONYMS = {
 MKT_COL: ["marketboard", "market", "board", "market_board", "площадка", "режим", "mode", "рынок"],
}


@dataclass(frozen=True)
class UnoKey:
	marketboard: str
	type_: str
	ticker: str
	date: str
	price_dec: Decimal


def _normalize_str(v: str) -> str:
	return (v or "").strip()


def _read_inout_csv(path: Path):
	delim = sniff_delimiter(path)
	rows = []
	with_board = False
	with path.open("r", encoding="utf-8-sig", newline="") as f:
		reader = csv.DictReader(f, delimiter=delim)
		header = [h or "" for h in (reader.fieldnames or [])]
		col_map = _resolve_columns(header, _INOUT_REQUIRED_SYNONYMS, _INOUT_OPTIONAL_SYNONYMS)
		with_board = MKT_COL in col_map
		for i, r in enumerate(reader):
			def get(canon: str) -> str:
				name = col_map.get(canon)
				return _normalize_str(r.get(name, "")) if name else ""
			rows.append(
			 {
			  "MARKETBOARD": get(MKT_COL) if with_board else "",
			  "TYPE": get("TYPE"),
			  "TICKER": get("TICKER"),
			  "VOLUME": _to_int_loose(get("VOLUME")),
			  "DATE_BUY": get("DATE_BUY"),
			  "PRICE_BUY": get("PRICE_BUY"),
			  "DATE_SELL": get("DATE_SELL"),
			  "PRICE_SELL": get("PRICE_SELL"),
			  "_ord": i,
			 }
			)
	return rows, with_board


def _read_inout_xlsx(path: Path):
	try:
		from openpyxl import load_workbook
	except Exception as e:
		raise RuntimeError("Package openpyxl is required: pip install openpyxl") from e

	wb = load_workbook(path, data_only=True)
	ws = wb.active
	header = []
	for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
		header.append(str(cell.value).strip() if cell.value is not None else "")

	with_board = MKT_COL in header
	col_map = _resolve_columns(header, _INOUT_REQUIRED_SYNONYMS, _INOUT_OPTIONAL_SYNONYMS)
	with_board = MKT_COL in col_map
	idx = {name: header.index(name) for name in header if name}
	rows = []
	for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
		if not row or all(v is None or str(v).strip() == "" for v in row):
			continue

		def get(canon: str) -> str:
			name = col_map.get(canon)
			j = idx.get(name) if name else None
			if j is None or j >= len(row):
				return ""
			v = row[j]
			if v is None:
				return ""

			if isinstance(v, (int, float, Decimal)):
				return str(v)
			if hasattr(v, "strftime"):
				try:
					return v.strftime(DATE_FMT)
				except Exception:
					return str(v)
			return str(v)

		vol_raw = get("VOLUME")
		rows.append(
		 {
		  "MARKETBOARD": _normalize_str(get(MKT_COL)) if with_board else "",
		  "TYPE": _normalize_str(get("TYPE")),
		  "TICKER": _normalize_str(get("TICKER")),
		  "VOLUME": _to_int_loose(vol_raw),
		  "DATE_BUY": _normalize_str(get("DATE_BUY")),
		  "PRICE_BUY": _normalize_str(get("PRICE_BUY")),
		  "DATE_SELL": _normalize_str(get("DATE_SELL")),
		  "PRICE_SELL": _normalize_str(get("PRICE_SELL")),
		  "_ord": i,
		 }
		)
	return rows, with_board


def read_inout(path: Path):
	suf = path.suffix.lower()
	if suf in (".csv", ".tsv", ".txt"):
		return _read_inout_csv(path)
	if suf == ".xlsx":
		return _read_inout_xlsx(path)
	raise ValueError("Supported input formats: .csv/.tsv/.xlsx")


def inout_to_uno(inout_rows: list[dict], with_board: bool):
	agg_qty = defaultdict(int)
	agg_frac = defaultdict(int)
	keep_type = {}
	keep_ticker = {}
	keep_board = {}

	def add_trade(marketboard: str, typ: str, ticker: str, date_s: str, price_raw: str, signed_qty: int):
		if not date_s or not price_raw or signed_qty == 0:
			return
		price_dec, _shown, frac_len = parse_price_keep_format(price_raw)
		key = UnoKey(marketboard=marketboard, type_=typ, ticker=ticker, date=date_s, price_dec=price_dec)
		agg_qty[key] += int(signed_qty)
		agg_frac[key] = max(int(agg_frac[key] or 0), int(frac_len or 0))
		keep_type[key] = typ
		keep_ticker[key] = ticker
		keep_board[key] = marketboard

	for r in inout_rows:
		marketboard = r.get("MARKETBOARD", "") if with_board else ""
		typ = r.get("TYPE", "")
		ticker = r.get("TICKER", "")
		vol = int(r.get("VOLUME", 0) or 0)
		if vol == 0:
			continue


		add_trade(marketboard, typ, ticker, r.get("DATE_BUY", ""), r.get("PRICE_BUY", ""), +vol)

		add_trade(marketboard, typ, ticker, r.get("DATE_SELL", ""), r.get("PRICE_SELL", ""), -vol)


	out = []
	for key, qty in agg_qty.items():
		if qty == 0:
			continue

		date_txt = key.date
		try:
			date_txt = parse_date(date_txt).strftime(DATE_FMT)
		except Exception:
			pass
		out.append(
		 {
		  "MARKETBOARD": keep_board.get(key, ""),
		  "TYPE": keep_type.get(key, key.type_),
		  "TICKER": keep_ticker.get(key, key.ticker),
		  "DATE": date_txt,
		  "VOLUME": int(qty),
		  "PRICE_DEC": key.price_dec,
		  "PRICE_FRAC": agg_frac.get(key, 0),
		 }
		)

	def sort_key(r):
		try:
			d = parse_date(r.get("DATE", ""))
		except Exception:
			d = datetime.max.date()

		side_rank = 0 if int(r.get("VOLUME", 0)) > 0 else 1
		return (
		 d,
		 r.get("TICKER", ""),
		 r.get("MARKETBOARD", "") if with_board else "",
		 r.get("PRICE_DEC"),
		 side_rank,
		)

	out.sort(key=sort_key)
	return out


def write_uno_csv(rows: list[dict], out_path: Path, with_board: bool):
	header = ["TYPE", "TICKER", "DATE", "VOLUME", "PRICE"]
	if with_board:
		header = [MKT_COL] + header

	with out_path.open("w", encoding="utf-8-sig", newline="") as f:
		w = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
		w.writerow(header)
		for r in rows:
			price_txt = format_num_for_csv(r.get("PRICE_DEC"), r.get("PRICE_FRAC", 0))
			if with_board:
				w.writerow([r.get("MARKETBOARD", ""), r.get("TYPE", ""), r.get("TICKER", ""), r.get("DATE", ""), r.get("VOLUME", 0), price_txt])
			else:
				w.writerow([r.get("TYPE", ""), r.get("TICKER", ""), r.get("DATE", ""), r.get("VOLUME", 0), price_txt])


def write_uno_xlsx(rows: list[dict], out_path: Path, with_board: bool):
	try:
		from openpyxl import Workbook
	except Exception as e:
		raise RuntimeError("Package openpyxl is required: pip install openpyxl") from e

	header = ["TYPE", "TICKER", "DATE", "VOLUME", "PRICE"]
	if with_board:
		header = [MKT_COL] + header

	wb = Workbook()
	ws = wb.active
	ws.title = "ONES"
	ws.append(header)

	for r in rows:
		price_dec = r.get("PRICE_DEC")
		if with_board:
			ws.append([r.get("MARKETBOARD", ""), r.get("TYPE", ""), r.get("TICKER", ""), r.get("DATE", ""), int(r.get("VOLUME", 0) or 0), price_dec])
		else:
			ws.append([r.get("TYPE", ""), r.get("TICKER", ""), r.get("DATE", ""), int(r.get("VOLUME", 0) or 0), price_dec])


	price_col = 6 if with_board else 5
	for i in range(2, len(rows) + 2):
		frac = int(rows[i - 2].get("PRICE_FRAC", 0) or 0)
		fmt = "#,##0" + (("." + "0" * frac) if frac > 0 else "")
		ws.cell(row=i, column=price_col).number_format = fmt

	wb.save(out_path)


def write_uno(rows: list[dict], out_path: Path, with_board: bool):
	suf = out_path.suffix.lower()
	if suf == ".csv":
		write_uno_csv(rows, out_path, with_board)
		return
	if suf == ".xlsx":
		write_uno_xlsx(rows, out_path, with_board)
		return
	raise ValueError("Specify output file .csv or .xlsx")


def print_to_stdout(rows: list[dict], with_board: bool):
	header = ["TYPE", "TICKER", "DATE", "VOLUME", "PRICE"]
	if with_board:
		header = [MKT_COL] + header
	print(", ".join(header))
	for r in rows:
		price_txt = format_num_for_csv(r.get("PRICE_DEC"), r.get("PRICE_FRAC", 0))
		if with_board:
			print(", ".join([str(r.get("MARKETBOARD", "")), str(r.get("TYPE", "")), str(r.get("TICKER", "")), str(r.get("DATE", "")), str(r.get("VOLUME", 0)), price_txt]))
		else:
			print(", ".join([str(r.get("TYPE", "")), str(r.get("TICKER", "")), str(r.get("DATE", "")), str(r.get("VOLUME", 0)), price_txt]))


def main():
	def _ask_path(prompt: str) -> Path:
		s = input(prompt).strip()
		if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
			s = s[1:-1].strip()
		return Path(s)

	if len(sys.argv) >= 3:
		inp = Path(sys.argv[1])
		outp = Path(sys.argv[2])
	else:
		print("Specify file paths.")
		print("Supported formats: input .csv/.tsv/.xlsx, output .csv/.xlsx")
		inp = _ask_path("INPUT TWOS (INOUT) file: ")
		outp = _ask_path("OUTPUT ONES (UNO) file: ")
		if not str(inp).strip() or not str(outp).strip():
			print("Empty path. Usage:\n  python feed_match.py INPUT.(csv|tsv|xlsx) OUTPUT.(csv|xlsx)")
			sys.exit(1)

	inout_rows, with_board = read_inout(inp)
	uno_rows = inout_to_uno(inout_rows, with_board=with_board)
	write_uno(uno_rows, outp, with_board=with_board)
	print_to_stdout(uno_rows, with_board=with_board)
	print(f"\nFile saved: {outp}  |  rows: {len(uno_rows)}")


if __name__ == "__main__":
	main()



read_twos = read_inout
twos_to_ones = inout_to_uno
write_ones_csv = write_uno_csv
write_ones_xlsx = write_uno_xlsx
write_ones = write_uno

