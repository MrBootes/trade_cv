#python3
# -*- coding: utf-8 -*-

import csv, re
import difflib
from collections import defaultdict, deque
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, getcontext, InvalidOperation
from pathlib import Path
import sys

getcontext().prec = 50

IN_COLS = ["TYPE", "TICKER", "DATE", "VOLUME", "PRICE"]
OUT_COLS = ["TYPE", "TICKER", "VOLUME", "DATE_BUY", "PRICE_BUY", "DATE_SELL", "PRICE_SELL"]
COMM_COL = "COMMISSION"
OUT_COLS_WITH_COMM = ["TYPE", "TICKER", "VOLUME", "DATE_BUY", "PRICE_BUY", "BUY_COMMISSION", "DATE_SELL", "PRICE_SELL", "SELL_COMMISSION"]
MKT_COL = "MARKETBOARD"
OUT_COLS_WITH_BOARD = ["MARKETBOARD", "TYPE", "TICKER", "VOLUME", "DATE_BUY", "PRICE_BUY", "DATE_SELL", "PRICE_SELL"]
OUT_COLS_WITH_COMM_AND_BOARD = ["MARKETBOARD", "TYPE", "TICKER", "VOLUME", "DATE_BUY", "PRICE_BUY", "BUY_COMMISSION", "DATE_SELL", "PRICE_SELL", "SELL_COMMISSION"]
DATE_FMT = "%d.%m.%Y"

def parse_date(s: str):
    txt = (s or "").strip()
    if not txt:
        raise ValueError("empty date")
    for fmt in (DATE_FMT, "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(txt, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(txt.replace("Z", "+00:00")).date()
    except Exception:
        pass
    raise ValueError(f"invalid date: '{s}'")


def _norm_col_name(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().casefold()
    for ch in ("\u00A0", "\u202F", "\u2009", " ", "\t", "\n", "\r"):
        s = s.replace(ch, "")
    s = "".join(c for c in s if c.isalnum())
    return s


def _to_int_loose(v) -> int:
    if v is None:
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, Decimal):
        return int(v)
    txt = str(v).strip()
    if not txt:
        return 0
    txt = strip_space_digits(txt)
    txt = txt.replace(",", "")
    try:
        return int(txt)
    except Exception:
        try:
            return int(float(txt))
        except Exception:
            return 0


def _resolve_columns(
    header: list[str],
    required: dict[str, list[str]],
    optional: dict[str, list[str]] | None = None,
    *,
    fuzzy_threshold: float = 0.86,
):
    optional = optional or {}

    norm_to_orig: dict[str, str] = {}
    for h in header:
        n = _norm_col_name(h)
        if n and n not in norm_to_orig:
            norm_to_orig[n] = h

    def candidates_for(canon: str) -> list[str]:
        alts = [canon] + list(required.get(canon, [])) + list(optional.get(canon, []))
        seen = set()
        out = []
        for a in alts:
            n = _norm_col_name(a)
            if n and n not in seen:
                seen.add(n)
                out.append(n)
        return out

    resolved: dict[str, str] = {}
    missing: list[str] = []
    all_norm_headers = list(norm_to_orig.keys())

    for canon in required.keys():
        found = None
        for alt_norm in candidates_for(canon):
            if alt_norm in norm_to_orig:
                found = norm_to_orig[alt_norm]
                break
        if found is None and all_norm_headers:
            best_norm = None
            best_score = 0.0
            acceptable = candidates_for(canon)
            for hn in all_norm_headers:
                for an in acceptable:
                    sc = difflib.SequenceMatcher(a=hn, b=an).ratio()
                    if sc > best_score:
                        best_score = sc
                        best_norm = hn
            if best_norm is not None and best_score >= fuzzy_threshold:
                found = norm_to_orig[best_norm]
        if found is None:
            missing.append(canon)
        else:
            resolved[canon] = found

    for canon in optional.keys():
        for alt_norm in candidates_for(canon):
            if alt_norm in norm_to_orig:
                resolved[canon] = norm_to_orig[alt_norm]
                break

    if missing:
        avail = [h for h in header if (h or "").strip()]
        suggestions = {}
        for canon in missing:
            expected_norms = candidates_for(canon)
            pairs = []
            for h in avail:
                hn = _norm_col_name(h)
                if not hn:
                    continue
                best = max((difflib.SequenceMatcher(a=hn, b=en).ratio() for en in expected_norms), default=0.0)
                pairs.append((best, h))
            pairs.sort(reverse=True)
            suggestions[canon] = [h for sc, h in pairs[:5] if sc >= 0.55]

        msg_lines = [
            f"Missing required columns: {missing}.",
            f"Found columns: {avail}",
        ]
        for canon in missing:
            if suggestions.get(canon):
                msg_lines.append(f"Suggestions for '{canon}': {suggestions[canon]}")
        raise ValueError("\n".join(msg_lines))

    return resolved


_UNO_REQUIRED_SYNONYMS = {
    "TYPE": ["type", "typ", "assettype", "вид", "тип"],
    "TICKER": ["ticker", "symbol", "security", "sec", "тикер", "тик", "бумага", "инструмент"],
    "DATE": ["date", "dt", "trade_date", "дата", "датасделки", "дата сделки"],
    "VOLUME": ["volume", "qty", "quantity", "amount", "count", "количество", "колво", "кол-во", "шт", "units"],
    "PRICE": ["price", "px", "rate", "цена", "курс"],
}


_UNO_OPTIONAL_SYNONYMS = {
    COMM_COL: ["commission", "fee", "comm", "комиссия", "сбор"],
    MKT_COL: ["marketboard", "market", "board", "market_board", "площадка", "режим", "mode", "рынок"],
}

# --- PRICE PARSER (preserve fractional length) ---
PRICE_RE = re.compile(r"""
    ^\s*
    (?P<sign>[+-]?)\s*
    (?P<int>[0-9\u00A0\u202F\u2009\s]*[0-9])?
    (?:[,.](?P<frac>[0-9]+))?
    .*?$
""", re.X)

def strip_space_digits(s: str) -> str:
    return s.replace("\u00A0","").replace("\u202F","").replace("\u2009","").replace(" ","")

def group_thousands(int_digits: str) -> str:
    s = int_digits
    out = []
    while s:
        out.append(s[-3:])
        s = s[:-3]
    return " ".join(reversed(out)) if out else "0"

def parse_price_keep_format(raw: str):
    """
    -> (Decimal value, printable_string_with_<currency>, frac_len)
    no rounding, frac_len = number of digits after decimal point in input
    """
    txt = (raw or "").replace("руб.", "").replace("р.", "").replace("₽", "").replace(" ", "")
    txt = txt.replace("$", "").replace("€", "").replace("£", "").replace("¥", "").replace("usd", "")
    txt = txt.replace("eur", "").replace("gbp", "").replace("jpy", "").strip()
    m = PRICE_RE.match(txt)
    if not m:
        raise ValueError(f"invalid price: '{raw}'")
    sign = m.group("sign") or ""
    int_part = strip_space_digits(m.group("int") or "0")
    frac_part = m.group("frac")  # can be None
    if not int_part.isdigit():
        raise ValueError(f"invalid integer part of price: '{raw}'")

    dec_str = f"{'-' if sign=='-' else ''}{int_part}"
    frac_len = 0
    if frac_part is not None:
        dec_str += "." + frac_part
        frac_len = len(frac_part)
    try:
        dec_val = Decimal(dec_str)
    except InvalidOperation as e:
        raise ValueError(f"failed to parse price: '{raw}' -> '{dec_str}'") from e

    shown_int = group_thousands(int_part.lstrip("0") or "0")
    if frac_part is None:
        shown = f"{shown_int} ₽"
    else:
        shown = f"{shown_int},{frac_part} ₽"
    if sign == "-":
        shown = "-" + shown
    return dec_val, shown, frac_len

def format_money_shown(dec: Decimal | None, frac_len: int) -> str:
    """Format Decimal as a readable money string with ₽, preserving frac_len."""
    if dec is None:
        return ""
    sign = "-" if dec < 0 else ""
    dec_abs = -dec if dec < 0 else dec
    # Build exact string with requested fractional digits
    if frac_len <= 0:
        q = Decimal('1')
        s = f"{dec_abs.quantize(q)}"
        int_part, frac_part = s, None
    else:
        q = Decimal(1).scaleb(-frac_len)
        s = f"{dec_abs.quantize(q)}"
        if '.' in s:
            int_part, frac_part = s.split('.', 1)
        else:
            int_part, frac_part = s, "0" * frac_len

    shown_int = group_thousands((int_part or "0").lstrip("0") or "0")
    if frac_part is None:
        shown = f"{shown_int} ₽"
    else:
        shown = f"{shown_int},{frac_part} ₽"
    return sign + shown

def money_for_qty(unit_dec: Decimal | None, frac_len: int, qty: int):
    """Compute money amount for qty using per-unit commission, preserving frac_len."""
    if unit_dec is None:
        return None, "", 0
    frac_len = int(frac_len or 0)
    q = Decimal('1') if frac_len <= 0 else Decimal(1).scaleb(-frac_len)
    amt = (unit_dec * Decimal(qty)).quantize(q)
    return amt, format_money_shown(amt, frac_len), frac_len

def format_num_for_csv(dec: Decimal | None, frac_len: int) -> str:
    """Number as text for CSV: no symbols/spaces, with exactly frac_len digits."""
    if dec is None:
        return ""
    # Format using quantize, but in CSV this will remain a number-text
    if frac_len <= 0:
        return f"{dec.quantize(Decimal('1'))}"
    q = Decimal(1).scaleb(-frac_len)  # 10^-frac_len
    return f"{dec.quantize(q)}"

@dataclass
class Leg:
    date: datetime
    price_dec: Decimal
    price_str: str
    frac_len: int
    qty: int
    marketboard: str = ""
    comm_unit_dec: Decimal | None = None
    comm_frac: int = 0

@dataclass
class PairRow:
    type_: str
    ticker: str
    marketboard: str
    volume: int
    date_buy: str
    price_buy_dec: Decimal | None
    price_buy_str: str
    price_buy_frac: int
    date_sell: str
    price_sell_dec: Decimal | None
    price_sell_str: str
    price_sell_frac: int
    buy_comm_dec: Decimal | None = None
    buy_comm_str: str = ""
    buy_comm_frac: int = 0
    sell_comm_dec: Decimal | None = None
    sell_comm_str: str = ""
    sell_comm_frac: int = 0

def sniff_delimiter(path: Path) -> str:
    text = path.read_text(encoding="utf-8-sig", errors="ignore")
    first_lines = [l for l in text.splitlines()[:5] if l.strip()]
    candidates = [';', '\t', ',']
    best, best_cols = ',', 0
    for cand in candidates:
        cols = min((len(l.split(cand)) for l in first_lines), default=0)
        if cols > best_cols:
            best_cols, best = cols, cand
    try:
        sniff = csv.Sniffer().sniff(text[:4096], delimiters=";,\t")
        sn_cols = len(first_lines[0].split(sniff.delimiter)) if first_lines else 0
        if sn_cols >= best_cols:
            return sniff.delimiter
    except Exception:
        pass
    return best

_XLSX_DEC_RE = re.compile(r"\.(?P<digits>[0#]+)")


def _excel_frac_len(fmt: str) -> int:
    """Best-effort: infer decimal digits from Excel number format like '#,##0.00'."""
    if not fmt:
        return 0
    m = _XLSX_DEC_RE.search(str(fmt))
    if not m:
        return 0
    return len(m.group("digits") or "")


def _excel_number_to_money(value, frac_len: int):
    """Convert Excel numeric (int/float/Decimal) to (Decimal, shown_str, frac_len)."""
    if value is None or value == "":
        raise ValueError("empty numeric value")
    if isinstance(value, Decimal):
        dec_val = value
    elif isinstance(value, int):
        dec_val = Decimal(value)
    elif isinstance(value, float):
        # avoid binary float artifacts as much as possible
        dec_val = Decimal(str(value))
    else:
        return parse_price_keep_format(str(value))

    frac_len = int(frac_len or 0)
    q = Decimal('1') if frac_len <= 0 else Decimal(1).scaleb(-frac_len)
    dec_val = dec_val.quantize(q)
    return dec_val, format_money_shown(dec_val, frac_len), frac_len


def _read_trades_xlsx(path: Path):
    """Read UNO trades from .xlsx (active sheet)."""
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("Package openpyxl is required: pip install openpyxl") from e

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1), [])
    header = [(str(c.value).strip() if c.value is not None else "") for c in header_cells]
    col_map = _resolve_columns(header, _UNO_REQUIRED_SYNONYMS, _UNO_OPTIONAL_SYNONYMS)
    has_commission = COMM_COL in col_map
    has_marketboard = MKT_COL in col_map

    idx = {name: header.index(name) for name in header if name}
    rows = []

    for i, row_cells in enumerate(ws.iter_rows(min_row=2)):
        if not row_cells or all(c.value is None or str(c.value).strip() == "" for c in row_cells):
            continue

        def cell(canon: str):
            name = col_map.get(canon)
            j = idx.get(name) if name else None
            return row_cells[j] if j is not None and j < len(row_cells) else None

        type_cell = cell("TYPE")
        tick_cell = cell("TICKER")
        date_cell = cell("DATE")
        vol_cell = cell("VOLUME")
        price_cell = cell("PRICE")

        typ = (str(type_cell.value).strip() if type_cell and type_cell.value is not None else "")
        ticker = (str(tick_cell.value).strip() if tick_cell and tick_cell.value is not None else "")

        # DATE: allow native Excel date/datetime or text dd.mm.yyyy
        date_v = date_cell.value if date_cell else ""
        if hasattr(date_v, "year") and hasattr(date_v, "month") and hasattr(date_v, "day"):
            dt = date_v.date() if hasattr(date_v, "date") else date_v
        else:
            dt = parse_date(str(date_v))

        # VOLUME
        vol_v = vol_cell.value if vol_cell else 0
        vol = _to_int_loose(vol_v)

        # PRICE
        price_v = price_cell.value if price_cell else ""
        if isinstance(price_v, (int, float, Decimal)):
            pf = _excel_frac_len(getattr(price_cell, "number_format", ""))
            p_dec, p_str, p_frac = _excel_number_to_money(price_v, pf)
        else:
            p_dec, p_str, p_frac = parse_price_keep_format(str(price_v))

        # COMMISSION (optional)
        c_dec = c_str = None
        c_shown = ""
        c_frac = 0
        if has_commission:
            comm_cell = cell(COMM_COL)
            comm_v = comm_cell.value if comm_cell else ""
            if comm_v is not None and str(comm_v).strip() != "":
                if isinstance(comm_v, (int, float, Decimal)):
                    cf = _excel_frac_len(getattr(comm_cell, "number_format", ""))
                    c_dec, c_shown, c_frac = _excel_number_to_money(comm_v, cf)
                else:
                    c_dec, c_shown, c_frac = parse_price_keep_format(str(comm_v))

        mb = ""
        if has_marketboard:
            mb_cell = cell(MKT_COL)
            mb = (str(mb_cell.value).strip() if mb_cell and mb_cell.value is not None else "")

        rows.append({
            "TYPE": typ,
            "TICKER": ticker,
            "DATE": dt,
            "VOLUME": vol,
            "PRICE_DEC": p_dec,
            "PRICE_STR": p_str,
            "PRICE_FRAC": p_frac,
            "COMM_DEC": c_dec,
            "COMM_STR": c_shown,
            "COMM_FRAC": c_frac,
            "MARKETBOARD": mb,
            "_ord": i
        })

    rows.sort(key=lambda x: (x["TICKER"], x["DATE"], x["_ord"]))
    return rows, has_commission, has_marketboard

def read_trades(path: Path):
    suf = path.suffix.lower()
    if suf == ".xlsx":
        return _read_trades_xlsx(path)
    if suf not in (".csv", ".tsv", ".txt"):
        raise ValueError(f"Unsupported input format: {suf}. Expected .csv/.tsv/.xlsx")
    delim = sniff_delimiter(path)
    rows = []
    has_commission = False
    has_marketboard = False
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delim)
        header = [h or "" for h in (reader.fieldnames or [])]
        col_map = _resolve_columns(header, _UNO_REQUIRED_SYNONYMS, _UNO_OPTIONAL_SYNONYMS)
        has_commission = COMM_COL in col_map
        has_marketboard = MKT_COL in col_map
        for i, r in enumerate(reader):
            def get(canon: str) -> str:
                name = col_map.get(canon)
                return (r.get(name, "") if name else "")

            p_dec, p_str, frac_len = parse_price_keep_format(get("PRICE"))
            c_dec = c_str = None
            c_shown = ""
            c_frac = 0
            if has_commission:
                raw_c = (get(COMM_COL) or "").strip()
                if raw_c:
                    c_dec, c_shown, c_frac = parse_price_keep_format(raw_c)
            mb = (get(MKT_COL) or "").strip() if has_marketboard else ""
            rows.append({
                "TYPE": (get("TYPE") or "").strip(),
                "TICKER": (get("TICKER") or "").strip(),
                "DATE": parse_date(get("DATE")),
                "VOLUME": _to_int_loose(get("VOLUME")),
                "PRICE_DEC": p_dec,
                "PRICE_STR": p_str,
                "PRICE_FRAC": frac_len,
                "COMM_DEC": c_dec,
                "COMM_STR": c_shown,
                "COMM_FRAC": c_frac,
                "MARKETBOARD": mb,
                "_ord": i
            })
    rows.sort(key=lambda x: (x["TICKER"], x["DATE"], x["_ord"]))
    return rows, has_commission, has_marketboard

def fifo_match(rows, has_commission: bool = False, has_marketboard: bool = False):
    queues = defaultdict(lambda: {"buys": deque(), "sells": deque(), "type": None})
    matched: list[PairRow] = []

    for r in rows:
        ticker, typ, dt, vol = r["TICKER"], r["TYPE"], r["DATE"], r["VOLUME"]
        p_dec, p_str, p_frac = r["PRICE_DEC"], r["PRICE_STR"], r["PRICE_FRAC"]
        c_dec, c_str, c_frac = r.get("COMM_DEC"), r.get("COMM_STR", ""), r.get("COMM_FRAC", 0)
        mb = r.get("MARKETBOARD", "") or ""
        trade_comm_unit = None
        if c_dec is not None and vol != 0:
            trade_comm_unit = c_dec / Decimal(abs(vol))
        if queues[ticker]["type"] is None:
            queues[ticker]["type"] = typ
        if vol == 0:
            continue

        if vol > 0:
            qty = vol
            sells = queues[ticker]["sells"]
            while qty > 0 and sells:
                sh = sells[0]
                m = min(qty, sh.qty)
                out_mb = mb or sh.marketboard
                buy_comm_dec, buy_comm_str, buy_comm_frac = money_for_qty(trade_comm_unit, c_frac, m)
                sell_comm_dec, sell_comm_str, sell_comm_frac = money_for_qty(sh.comm_unit_dec, sh.comm_frac, m)
                matched.append(PairRow(
                    typ, ticker, out_mb, m,
                    dt.strftime(DATE_FMT), p_dec, p_str, p_frac,
                    sh.date.strftime(DATE_FMT), sh.price_dec, sh.price_str, sh.frac_len
                    ,
                    buy_comm_dec, buy_comm_str, buy_comm_frac,
                    sell_comm_dec, sell_comm_str, sell_comm_frac
                ))
                qty -= m; sh.qty -= m
                if sh.qty == 0: sells.popleft()
            if qty > 0:
                lg = Leg(dt, p_dec, p_str, p_frac, qty, mb, trade_comm_unit, c_frac)
                queues[ticker]["buys"].append(lg)
        else:
            qty = -vol
            buys = queues[ticker]["buys"]
            while qty > 0 and buys:
                lg = buys[0]
                m = min(qty, lg.qty)
                out_mb = mb or lg.marketboard
                buy_comm_dec, buy_comm_str, buy_comm_frac = money_for_qty(lg.comm_unit_dec, lg.comm_frac, m)
                sell_comm_dec, sell_comm_str, sell_comm_frac = money_for_qty(trade_comm_unit, c_frac, m)
                matched.append(PairRow(
                    typ, ticker, out_mb, m,
                    lg.date.strftime(DATE_FMT), lg.price_dec, lg.price_str, lg.frac_len,
                    dt.strftime(DATE_FMT), p_dec, p_str, p_frac
                    ,
                    buy_comm_dec, buy_comm_str, buy_comm_frac,
                    sell_comm_dec, sell_comm_str, sell_comm_frac
                ))
                qty -= m; lg.qty -= m
                if lg.qty == 0: buys.popleft()
            if qty > 0:
                sh = Leg(dt, p_dec, p_str, p_frac, qty, mb, trade_comm_unit, c_frac)
                queues[ticker]["sells"].append(sh)

    for ticker, b in queues.items():
        typ = b["type"] or "futures"
        for leg in b["buys"]:
            bc_dec, bc_str, bc_frac = money_for_qty(leg.comm_unit_dec, leg.comm_frac, leg.qty)
            matched.append(PairRow(typ, ticker, leg.marketboard, leg.qty,
                                   leg.date.strftime(DATE_FMT), leg.price_dec, leg.price_str, leg.frac_len,
                                   "", None, "", 0,
                                   bc_dec, bc_str, bc_frac,
                                   None, "", 0))
        for leg in b["sells"]:
            sc_dec, sc_str, sc_frac = money_for_qty(leg.comm_unit_dec, leg.comm_frac, leg.qty)
            matched.append(PairRow(typ, ticker, leg.marketboard, leg.qty,
                                   "", None, "", 0,
                                   leg.date.strftime(DATE_FMT), leg.price_dec, leg.price_str, leg.frac_len,
                                   None, "", 0,
                                   sc_dec, sc_str, sc_frac))

    # Aggregation
    agg_vol = defaultdict(int)
    agg_bc = defaultdict(lambda: Decimal('0'))
    agg_sc = defaultdict(lambda: Decimal('0'))
    keep = {}
    for row in matched:
        key = (row.type_, row.ticker, row.marketboard, row.date_buy, row.price_buy_dec, row.date_sell, row.price_sell_dec)
        agg_vol[key] += row.volume
        if row.buy_comm_dec is not None:
            agg_bc[key] += row.buy_comm_dec
        if row.sell_comm_dec is not None:
            agg_sc[key] += row.sell_comm_dec
        if key not in keep:
            keep[key] = (
                row.price_buy_str, row.price_buy_frac,
                row.price_sell_str, row.price_sell_frac,
                row.buy_comm_frac, row.sell_comm_frac,
            )
        else:
            pb_str, pb_frac, ps_str, ps_frac, bc_frac, sc_frac = keep[key]
            keep[key] = (
                pb_str, pb_frac,
                ps_str, ps_frac,
                max(bc_frac or 0, row.buy_comm_frac or 0),
                max(sc_frac or 0, row.sell_comm_frac or 0),
            )

    out_rows = []
    for (typ, ticker, mb, db, pb_dec, ds, ps_dec), vol in agg_vol.items():
        pb_str, pb_frac, ps_str, ps_frac, bc_frac, sc_frac = keep[(typ, ticker, mb, db, pb_dec, ds, ps_dec)]
        bc_dec = agg_bc[(typ, ticker, mb, db, pb_dec, ds, ps_dec)] if (typ, ticker, mb, db, pb_dec, ds, ps_dec) in agg_bc else Decimal('0')
        sc_dec = agg_sc[(typ, ticker, mb, db, pb_dec, ds, ps_dec)] if (typ, ticker, mb, db, pb_dec, ds, ps_dec) in agg_sc else Decimal('0')
        out_rows.append({
            "TYPE": typ, "TICKER": ticker, "VOLUME": vol,
            "MARKETBOARD": mb,
            "DATE_BUY": db, "PRICE_BUY_STR": pb_str, "PRICE_BUY_NUM": pb_dec, "PRICE_BUY_FRAC": pb_frac,
            "DATE_SELL": ds, "PRICE_SELL_STR": ps_str, "PRICE_SELL_NUM": ps_dec, "PRICE_SELL_FRAC": ps_frac,
            "BUY_COMM_STR": format_money_shown(bc_dec, bc_frac) if bc_frac is not None else "",
            "BUY_COMM_NUM": bc_dec,
            "BUY_COMM_FRAC": bc_frac,
            "SELL_COMM_STR": format_money_shown(sc_dec, sc_frac) if sc_frac is not None else "",
            "SELL_COMM_NUM": sc_dec,
            "SELL_COMM_FRAC": sc_frac,
        })

    # Sort output
    def d(x): return datetime.strptime(x, DATE_FMT).date() if x else datetime.max.date()
    out_rows.sort(key=lambda r: (r["TICKER"], r.get("MARKETBOARD", ""), d(r["DATE_BUY"]), d(r["DATE_SELL"]), r["VOLUME"]))

    flat = []
    for r in out_rows:
        base = {
            "TYPE": r["TYPE"], "TICKER": r["TICKER"], "VOLUME": r["VOLUME"],
            "DATE_BUY": r["DATE_BUY"],
            "PRICE_BUY": r["PRICE_BUY_STR"],      # for console (pretty, with ₽)
            "DATE_SELL": r["DATE_SELL"],
            "PRICE_SELL": r["PRICE_SELL_STR"],    # for console
            # for recording file:
            "_PB_NUM": r["PRICE_BUY_NUM"], "_PB_FRAC": r["PRICE_BUY_FRAC"],
            "_PS_NUM": r["PRICE_SELL_NUM"], "_PS_FRAC": r["PRICE_SELL_FRAC"],
        }
        if has_marketboard:
            base.update({"MARKETBOARD": r.get("MARKETBOARD", "")})
        if has_commission:
            base.update({
                "BUY_COMMISSION": r.get("BUY_COMM_STR", ""),
                "SELL_COMMISSION": r.get("SELL_COMM_STR", ""),
                "_BC_NUM": r.get("BUY_COMM_NUM"), "_BC_FRAC": r.get("BUY_COMM_FRAC", 0),
                "_SC_NUM": r.get("SELL_COMM_NUM"), "_SC_FRAC": r.get("SELL_COMM_FRAC", 0),
            })
        flat.append(base)
    return flat

# --- writing files ---

def write_csv_numeric_money(rows, out_path: Path):
    # In CSV, write numbers without currency signs/spaces and without quotes; decimal point; preserve number of digits.
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        with_comm = bool(rows and ("_BC_NUM" in rows[0] or "_SC_NUM" in rows[0]))
        with_board = bool(rows and ("MARKETBOARD" in rows[0]))
        if with_comm and with_board:
            writer.writerow(OUT_COLS_WITH_COMM_AND_BOARD)
        elif with_comm:
            writer.writerow(OUT_COLS_WITH_COMM)
        elif with_board:
            writer.writerow(OUT_COLS_WITH_BOARD)
        else:
            writer.writerow(OUT_COLS)
        for r in rows:
            pb = format_num_for_csv(r["_PB_NUM"], r["_PB_FRAC"])
            ps = format_num_for_csv(r["_PS_NUM"], r["_PS_FRAC"])
            if with_comm:
                bc = format_num_for_csv(r.get("_BC_NUM"), r.get("_BC_FRAC", 0))
                sc = format_num_for_csv(r.get("_SC_NUM"), r.get("_SC_FRAC", 0))
                if with_board:
                    writer.writerow([r.get("MARKETBOARD", ""), r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], pb, bc, r["DATE_SELL"], ps, sc])
                else:
                    writer.writerow([r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], pb, bc, r["DATE_SELL"], ps, sc])
            else:
                if with_board:
                    writer.writerow([r.get("MARKETBOARD", ""), r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], pb, r["DATE_SELL"], ps])
                else:
                    writer.writerow([r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], pb, r["DATE_SELL"], ps])

def write_xlsx_money(rows, out_path: Path):
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("Package openpyxl is required: pip install openpyxl") from e
    wb = Workbook()
    ws = wb.active
    ws.title = "FIFO"
    with_comm = bool(rows and ("_BC_NUM" in rows[0] or "_SC_NUM" in rows[0]))
    with_board = bool(rows and ("MARKETBOARD" in rows[0]))
    if with_comm and with_board:
        ws.append(OUT_COLS_WITH_COMM_AND_BOARD)
    elif with_comm:
        ws.append(OUT_COLS_WITH_COMM)
    elif with_board:
        ws.append(OUT_COLS_WITH_BOARD)
    else:
        ws.append(OUT_COLS)
    for r in rows:
        if with_comm:
            if with_board:
                ws.append([r.get("MARKETBOARD", ""), r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], r["_PB_NUM"], r.get("_BC_NUM"), r["DATE_SELL"], r["_PS_NUM"], r.get("_SC_NUM")])
            else:
                ws.append([r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], r["_PB_NUM"], r.get("_BC_NUM"), r["DATE_SELL"], r["_PS_NUM"], r.get("_SC_NUM")])
        else:
            if with_board:
                ws.append([r.get("MARKETBOARD", ""), r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], r["_PB_NUM"], r["DATE_SELL"], r["_PS_NUM"]])
            else:
                ws.append([r["TYPE"], r["TICKER"], r["VOLUME"], r["DATE_BUY"], r["_PB_NUM"], r["DATE_SELL"], r["_PS_NUM"]])
    for row_i in range(2, len(rows)+2):
        pb_frac = rows[row_i-2]["_PB_FRAC"]
        ps_frac = rows[row_i-2]["_PS_FRAC"]
        fmt_pb = "#,##0" + (("." + "0"*pb_frac) if pb_frac > 0 else "")
        fmt_ps = "#,##0" + (("." + "0"*ps_frac) if ps_frac > 0 else "")
        if with_comm:
            bc_frac = rows[row_i-2].get("_BC_FRAC", 0)
            sc_frac = rows[row_i-2].get("_SC_FRAC", 0)
            fmt_bc = "#,##0" + (("." + "0"*bc_frac) if bc_frac > 0 else "")
            fmt_sc = "#,##0" + (("." + "0"*sc_frac) if sc_frac > 0 else "")
            if with_board:
                ws.cell(row=row_i, column=6).number_format = fmt_pb
                ws.cell(row=row_i, column=7).number_format = fmt_bc
                ws.cell(row=row_i, column=9).number_format = fmt_ps
                ws.cell(row=row_i, column=10).number_format = fmt_sc
            else:
                ws.cell(row=row_i, column=5).number_format = fmt_pb
                ws.cell(row=row_i, column=6).number_format = fmt_bc
                ws.cell(row=row_i, column=8).number_format = fmt_ps
                ws.cell(row=row_i, column=9).number_format = fmt_sc
        else:
            if with_board:
                ws.cell(row=row_i, column=6).number_format = fmt_pb
                ws.cell(row=row_i, column=8).number_format = fmt_ps
            else:
                ws.cell(row=row_i, column=5).number_format = fmt_pb
                ws.cell(row=row_i, column=7).number_format = fmt_ps
    wb.save(out_path)

def write_file(rows, out_path: Path):
    suf = out_path.suffix.lower()
    if suf == ".csv":
        write_csv_numeric_money(rows, out_path)
    elif suf == ".xlsx":
        write_xlsx_money(rows, out_path)
    else:
        raise ValueError("Specify output file .csv or .xlsx")


def print_to_stdout(rows):
    with_comm = bool(rows and ("BUY_COMMISSION" in rows[0] or "SELL_COMMISSION" in rows[0]))
    with_board = bool(rows and ("MARKETBOARD" in rows[0]))
    if with_comm and with_board:
        print(", ".join(OUT_COLS_WITH_COMM_AND_BOARD))
    elif with_comm:
        print(", ".join(OUT_COLS_WITH_COMM))
    elif with_board:
        print(", ".join(OUT_COLS_WITH_BOARD))
    else:
        print(", ".join(OUT_COLS))
    for r in rows:
        if with_comm:
            if with_board:
                print(", ".join([
                    str(r.get("MARKETBOARD", "")), str(r["TYPE"]), str(r["TICKER"]), str(r["VOLUME"]),
                    r["DATE_BUY"], r["PRICE_BUY"], r.get("BUY_COMMISSION", ""), r["DATE_SELL"], r["PRICE_SELL"], r.get("SELL_COMMISSION", "")
                ]))
            else:
                print(", ".join([
                    str(r["TYPE"]), str(r["TICKER"]), str(r["VOLUME"]),
                    r["DATE_BUY"], r["PRICE_BUY"], r.get("BUY_COMMISSION", ""), r["DATE_SELL"], r["PRICE_SELL"], r.get("SELL_COMMISSION", "")
                ]))
        else:
            if with_board:
                print(", ".join([
                    str(r.get("MARKETBOARD", "")), str(r["TYPE"]), str(r["TICKER"]), str(r["VOLUME"]),
                    r["DATE_BUY"], r["PRICE_BUY"], r["DATE_SELL"], r["PRICE_SELL"]
                ]))
            else:
                print(", ".join([
                    str(r["TYPE"]), str(r["TICKER"]), str(r["VOLUME"]),
                    r["DATE_BUY"], r["PRICE_BUY"], r["DATE_SELL"], r["PRICE_SELL"]
                ]))

def main():
    def _ask_path(prompt: str) -> Path:
        s = input(prompt).strip()
        if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
            s = s[1:-1].strip()
        return Path(s)

    if len(sys.argv) >= 3:
        inp = Path(sys.argv[1])
        outp = Path(sys.argv[2])
    else:
        print("Specify file paths.")
        print("Supported formats: input .csv/.tsv/.xlsx, output .csv/.xlsx")
        inp = _ask_path("INPUT file: ")
        outp = _ask_path("OUTPUT file: ")
        if not str(inp).strip() or not str(outp).strip():
            print("Empty path. Usage:\n  python fifo_match.py INPUT.(csv|tsv|xlsx) OUTPUT.(csv|xlsx)")
            sys.exit(1)
    trades, has_commission, has_marketboard = read_trades(inp)
    result = fifo_match(trades, has_commission=has_commission, has_marketboard=has_marketboard)
    write_file(result, outp)
    print_to_stdout(result)
    print(f"\nFile saved: {outp}  |  rows: {len(result)}")

if __name__ == "__main__":
    main()
